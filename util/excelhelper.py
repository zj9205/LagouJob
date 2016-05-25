import json
import os
import logging

from openpyxl import Workbook

from util import toolkit

logging.basicConfig(filename='info.log', level=logging.DEBUG)


def json_to_list(job_type_json_dir):
    job_type_lists = []  # 每一个特定职位的list
    for each_json in os.listdir(job_type_json_dir):
        with open(job_type_json_dir + '/' + each_json, 'r', encoding='utf-8') as f:
            for each_line in f.readlines():
                json_content = '{"joblist":' + each_line + '}'
                json_obj = json.loads(
                    json_content.replace("\'", '\"').replace('None', 'null').replace('False', 'false'),
                    encoding='utf-8')
                job_type_lists.append(json_obj)

    return job_type_lists


def write_excel(lists, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "职位信息"
    ws.cell(row=1, column=1).value = '发布时间'
    ws.cell(row=1, column=2).value = '工作时间'
    ws.cell(row=1, column=3).value = '职位名称'
    ws.cell(row=1, column=4).value = '职位ID'
    ws.cell(row=1, column=5).value = '公司ID'
    ws.cell(row=1, column=6).value = '职位类型'
    ws.cell(row=1, column=7).value = '公司名称'
    ws.cell(row=1, column=8).value = '所在城市'
    ws.cell(row=1, column=9).value = '文化程度'
    ws.cell(row=1, column=10).value = '所属行业'
    ws.cell(row=1, column=11).value = '融资阶段'
    ws.cell(row=1, column=12).value = '职位薪酬'
    ws.cell(row=1, column=13).value = '公司规模'
    ws.cell(row=1, column=14).value = '平均薪资'

    rownum = 2

    for each_item in lists:
        info_list = each_item.get('joblist')
        for each_job_info_obj in info_list:
            ws.cell(row=rownum, column=1).value = each_job_info_obj['formatCreateTime']
            ws.cell(row=rownum, column=2).value = each_job_info_obj['workYear']
            ws.cell(row=rownum, column=3).value = each_job_info_obj['positionName']
            ws.cell(row=rownum, column=4).value = each_job_info_obj['positionId']
            ws.cell(row=rownum, column=5).value = each_job_info_obj['companyId']
            ws.cell(row=rownum, column=6).value = each_job_info_obj['positionType']
            ws.cell(row=rownum, column=7).value = each_job_info_obj['companyName']
            ws.cell(row=rownum, column=8).value = each_job_info_obj['city']
            ws.cell(row=rownum, column=9).value = each_job_info_obj['education']
            ws.cell(row=rownum, column=10).value = each_job_info_obj['industryField']
            ws.cell(row=rownum, column=11).value = each_job_info_obj['financeStage']
            ws.cell(row=rownum, column=12).value = each_job_info_obj['salary']
            ws.cell(row=rownum, column=13).value = each_job_info_obj['companySize']
            ws.cell(row=rownum, column=14).value = toolkit.normalize(each_job_info_obj['salary'])
            rownum += 1
    wb.save('d:/' + filename + '.xlsx')
    logging.info('Excel生成成功!')


def process(json_file_path):
    if os.path.exists(json_file_path):
        dir_list = os.listdir(json_file_path)
        for each_dir in dir_list:
            print(json_file_path + os.path.sep + each_dir)
            lists = json_to_list(json_file_path + os.path.sep + each_dir)
            write_excel(lists, each_dir)


if __name__ == '__main__':
    logging.info('start generating Excel file...')
    process('D:/LagouJobInfo/lagou')
    logging.info('Done! Please check your result...')
