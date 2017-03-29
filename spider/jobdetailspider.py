import os

import requests
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def get_detail_info_byid(job_id, outputfolder):
    """分析每个职位页面详情"""

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, sdch',
        'Host': 'www.lagou.com',
        'Upgrade-Insecure-Requests': 1,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.116 Safari/537.36'
    }
    if os.path.exists(outputfolder) is not True:
        os.mkdir(outputfolder)

    req_url = 'http://www.lagou.com/jobs/' + job_id + '.html'

    response = requests.get(req_url, headers=headers)

    if response.status_code == 200:
        html = response.content
        html_soup = BeautifulSoup(html, 'html.parser')

        job_bt_soup = html_soup.find('dd', class_='job_bt')

        # 加载用户自定义词典
        # jieba.load_userdict('C:/Users/XuLu/PycharmProjects/LagouJob/userdict.txt')

        # 加载停用词
        # jieba.analyse.set_stop_words('C:/Users/XuLu/PycharmProjects/LagouJob/stopwords.txt')

        # 这里是为了防止请求已失效的职位数据
        if job_bt_soup is not None:
            str_txt = job_bt_soup.text
            print(str_txt)

            result_txt_path = outputfolder + os.sep + job_id + '.txt'
            with open(result_txt_path, 'wt', encoding='utf-8') as f:
                f.write(str_txt)
                f.flush()
                f.close()
                print('职位编号为' + job_id + ' 的数据写出完成...')
        time.sleep(2)
    elif response.status_code == 403:
        print('Sorry, your requests are forbidden!')
    else:
        print('Something is wrong during the request...')


def get_jobid_list(job_excel_file_path):
    """从Excel文件里得到每个职位的职位ID，用来做进一步分析"""

    job_id_list = []
    wb = load_workbook(job_excel_file_path)
    ws = wb['职位信息']
    for rowindex in range(2, len(ws.rows) + 1):
        job_id = ws.cell(row=rowindex, column=4).value
        job_id_list.append(str(job_id))

    return job_id_list


if __name__ == '__main__':
    job_id_list = get_jobid_list('D:/计算机视觉.xlsx')
    outputdir = 'd:/'

    for each_job_id in job_id_list:
        get_detail_info_byid(each_job_id, 'D:/LagouJobInfo/lagou/details/计算机视觉')
